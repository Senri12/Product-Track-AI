#!/usr/bin/env python3
"""
Скрипт для анализа результатов моделей и создания Excel отчета
с условным форматированием и анализом языка
"""

import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import numpy as np

# Цвета для условного форматирования
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

def load_dialog_data(model_name, prompt_id, dialog_id):
    """Загрузка данных диалога из JSON файла"""
    model_folder = model_name.replace(':', '_')
    dialog_file = f"model_outputs/model_outputs/{model_folder}/dialog_{prompt_id}_{dialog_id}.json"

    if os.path.exists(dialog_file):
        with open(dialog_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None

def analyze_language_correctness(dialog_data):
    """
    Анализ корректности языка в ответах модели
    Возвращает: (is_correct, language, issues)
    """
    if not dialog_data:
        return None, None, "Нет данных"

    issues = []
    detected_language = None

    # Проверяем ответы модели
    if 'messages' in dialog_data:
        for msg in dialog_data['messages']:
            if msg.get('role') == 'assistant':
                content = msg.get('content', '')

                # Определяем язык по содержимому
                has_cyrillic = bool(any(ord(c) >= 0x0400 and ord(c) <= 0x04FF for c in content))
                has_latin = bool(any(c.isalpha() and ord(c) < 128 for c in content))

                if has_cyrillic and not has_latin:
                    detected_language = 'RU'
                elif has_latin and not has_cyrillic:
                    detected_language = 'EN'
                elif has_cyrillic and has_latin:
                    detected_language = 'MIXED'
                else:
                    detected_language = 'UNKNOWN'

                # Проверяем наличие английского в русском тексте
                if has_cyrillic:
                    # Подсчет английских слов (простая эвристика)
                    words = content.split()
                    en_words = [w for w in words if w.isalpha() and all(ord(c) < 128 for c in w) and len(w) > 2]
                    if len(en_words) > 3:  # Больше 3 английских слов
                        issues.append(f"Обнаружено {len(en_words)} англ. слов в русском тексте")

    # Проверяем оценку языка из RAGChecker (если есть)
    rag_file = dialog_data.get('rag_evaluation_file', '')
    if rag_file and os.path.exists(rag_file):
        try:
            with open(rag_file, 'r', encoding='utf-8') as f:
                rag_data = json.load(f)
                # Здесь можно добавить анализ RAG метрик
        except:
            pass

    is_correct = detected_language == 'RU' and len(issues) == 0

    return is_correct, detected_language, "; ".join(issues) if issues else "OK"

def analyze_accuracy_factors(dialog_data, rating):
    """
    Анализ факторов, влияющих на точность
    """
    factors = []

    if not dialog_data:
        return "Нет данных"

    # Длина контекста
    if 'messages' in dialog_data:
        total_chars = sum(len(msg.get('content', '')) for msg in dialog_data['messages'])
        if total_chars > 5000:
            factors.append("Длинный контекст")
        elif total_chars < 500:
            factors.append("Короткий контекст")

    # Рейтинг
    if rating >= 4.5:
        factors.append("Высокий рейтинг")
    elif rating <= 2.5:
        factors.append("Низкий рейтинг")

    # Проверка наличия RAG оценки
    if 'rag_evaluation_file' in dialog_data:
        factors.append("Есть RAG оценка")

    return "; ".join(factors) if factors else "Норма"

def create_analysis_excel(input_csv, output_xlsx):
    """Создание Excel файла с анализом"""

    print("Загрузка CSV данных...")
    df = pd.read_csv(input_csv)

    # Добавляем колонки для анализа
    print("Анализ данных...")
    df['Язык корректен'] = None
    df['Обнаруженный язык'] = None
    df['Проблемы языка'] = None
    df['Факторы точности'] = None

    # Анализируем каждую строку
    for idx, row in df.iterrows():
        if idx % 100 == 0:
            print(f"Обработано {idx}/{len(df)} строк...")

        dialog_data = load_dialog_data(row['model_name'], row['system_prompt_id'], row['dialog_id'])

        # Анализ языка
        is_correct, lang, issues = analyze_language_correctness(dialog_data)
        df.at[idx, 'Язык корректен'] = 'ДА' if is_correct else 'НЕТ' if is_correct is not None else 'N/A'
        df.at[idx, 'Обнаруженный язык'] = lang if lang else 'N/A'
        df.at[idx, 'Проблемы языка'] = issues

        # Анализ факторов точности
        df.at[idx, 'Факторы точности'] = analyze_accuracy_factors(dialog_data, row['overall_rating'])

    print("Создание Excel файла...")

    # Создаем основной лист с данными
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Все данные', index=False)

        # Сводная таблица по моделям
        print("Создание сводки по моделям...")
        model_summary = df.groupby('model_name').agg({
            'overall_rating': ['mean', 'std', 'min', 'max', 'count'],
            'Язык корректен': lambda x: (x == 'ДА').sum(),
        }).round(2)
        model_summary.columns = ['Средний рейтинг', 'Std', 'Min', 'Max', 'Кол-во', 'Язык OK']
        model_summary['% корректного языка'] = (model_summary['Язык OK'] / model_summary['Кол-во'] * 100).round(1)
        model_summary.to_excel(writer, sheet_name='По моделям')

        # Сводная таблица по промптам
        print("Создание сводки по промптам...")
        prompt_summary = df.groupby('system_prompt_id').agg({
            'overall_rating': ['mean', 'std', 'min', 'max', 'count'],
            'Язык корректен': lambda x: (x == 'ДА').sum(),
        }).round(2)
        prompt_summary.columns = ['Средний рейтинг', 'Std', 'Min', 'Max', 'Кол-во', 'Язык OK']
        prompt_summary['% корректного языка'] = (prompt_summary['Язык OK'] / prompt_summary['Кол-во'] * 100).round(1)
        prompt_summary.to_excel(writer, sheet_name='По промптам')

        # Матрица: модели vs промпты
        print("Создание матрицы модели-промпты...")
        pivot_rating = df.pivot_table(
            values='overall_rating',
            index='model_name',
            columns='system_prompt_id',
            aggfunc='mean'
        ).round(2)
        pivot_rating.to_excel(writer, sheet_name='Матрица рейтингов')

        # Матрица корректности языка
        df_lang_numeric = df.copy()
        df_lang_numeric['Язык корректен_num'] = (df_lang_numeric['Язык корректен'] == 'ДА').astype(int)
        pivot_lang = df_lang_numeric.pivot_table(
            values='Язык корректен_num',
            index='model_name',
            columns='system_prompt_id',
            aggfunc='mean'
        ).round(2) * 100  # В процентах
        pivot_lang.to_excel(writer, sheet_name='Матрица языка %')

    print("Применение условного форматирования...")

    # Открываем файл для добавления форматирования
    wb = load_workbook(output_xlsx)

    # Форматирование листа "Все данные"
    ws_all = wb['Все данные']
    format_all_data_sheet(ws_all, df)

    # Форматирование листа "По моделям"
    ws_models = wb['По моделям']
    format_summary_sheet(ws_models, 'models')

    # Форматирование листа "По промптам"
    ws_prompts = wb['По промптам']
    format_summary_sheet(ws_prompts, 'prompts')

    # Форматирование матрицы рейтингов
    ws_matrix = wb['Матрица рейтингов']
    format_matrix_sheet(ws_matrix, is_percent=False)

    # Форматирование матрицы языка
    ws_lang = wb['Матрица языка %']
    format_matrix_sheet(ws_lang, is_percent=True)

    wb.save(output_xlsx)
    print(f"Готово! Файл сохранен: {output_xlsx}")

def format_all_data_sheet(ws, df):
    """Форматирование листа со всеми данными"""

    # Заголовки
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Найти колонки
    header_row = [cell.value for cell in ws[1]]
    rating_col = header_row.index('overall_rating') + 1 if 'overall_rating' in header_row else None
    lang_col = header_row.index('Язык корректен') + 1 if 'Язык корректен' in header_row else None

    # Условное форматирование для рейтинга (цветовая шкала)
    if rating_col:
        col_letter = get_column_letter(rating_col)
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=1, start_color='FFC7CE',
                mid_type='num', mid_value=3, mid_color='FFEB9C',
                end_type='num', end_value=5, end_color='C6EFCE'
            )
        )

    # Условное форматирование для языка
    if lang_col:
        col_letter = get_column_letter(lang_col)
        # Зеленый для "ДА"
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"ДА"'], fill=GREEN_FILL)
        )
        # Красный для "НЕТ"
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"НЕТ"'], fill=RED_FILL)
        )
        # Желтый для "N/A"
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"N/A"'], fill=YELLOW_FILL)
        )

    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Закрепить первую строку
    ws.freeze_panes = 'A2'

def format_summary_sheet(ws, sheet_type):
    """Форматирование сводных листов"""

    # Заголовки
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Найти колонку среднего рейтинга
    header_row = [cell.value for cell in ws[1]]
    rating_col = header_row.index('Средний рейтинг') + 1 if 'Средний рейтинг' in header_row else None
    percent_col = header_row.index('% корректного языка') + 1 if '% корректного языка' in header_row else None

    # Цветовая шкала для среднего рейтинга
    if rating_col:
        col_letter = get_column_letter(rating_col)
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=1, start_color='FFC7CE',
                mid_type='num', mid_value=3, mid_color='FFEB9C',
                end_type='num', end_value=5, end_color='C6EFCE'
            )
        )

    # Цветовая шкала для процента корректного языка
    if percent_col:
        col_letter = get_column_letter(percent_col)
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FFC7CE',
                mid_type='num', mid_value=50, mid_color='FFEB9C',
                end_type='num', end_value=100, end_color='C6EFCE'
            )
        )

    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'B2'

def format_matrix_sheet(ws, is_percent=False):
    """Форматирование матричных листов"""

    # Заголовки
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Первая колонка (названия моделей)
    for cell in ws['A']:
        if cell.row > 1:
            cell.fill = LIGHT_BLUE_FILL
            cell.font = Font(bold=True)

    # Применить цветовую шкалу ко всем данным
    if ws.max_row > 1 and ws.max_column > 1:
        start_col = get_column_letter(2)
        end_col = get_column_letter(ws.max_column)

        if is_percent:
            # Для процентов: 0-100
            ws.conditional_formatting.add(
                f'{start_col}2:{end_col}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='FFC7CE',
                    mid_type='num', mid_value=50, mid_color='FFEB9C',
                    end_type='num', end_value=100, end_color='C6EFCE'
                )
            )
        else:
            # Для рейтингов: 1-5
            ws.conditional_formatting.add(
                f'{start_col}2:{end_col}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=1, start_color='FFC7CE',
                    mid_type='num', mid_value=3, mid_color='FFEB9C',
                    end_type='num', end_value=5, end_color='C6EFCE'
                )
            )

    # Авто-ширина
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'B2'

if __name__ == '__main__':
    input_csv = 'model_outputs/overall_report.csv'
    output_xlsx = 'analysis_report.xlsx'

    create_analysis_excel(input_csv, output_xlsx)
