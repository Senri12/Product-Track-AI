#!/usr/bin/env python3
"""
Скрипт для перекрестного сравнения двух версий данных RAGChecker
Version 1: английские промпты
Version 2: русские промпты
"""

import pandas as pd
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
import numpy as np

# Цвета для условного форматирования
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
LIGHT_GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
IMPROVEMENT_FILL = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
DEGRADATION_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

def load_prompts(version_path):
    """Загрузка промптов из JSON файла"""
    prompts_file = Path(version_path) / "system_prompts.json"

    if not prompts_file.exists():
        print(f"Файл не найден: {prompts_file}")
        return {}

    with open(prompts_file, 'r', encoding='utf-8') as f:
        prompts_data = json.load(f)

    # Преобразуем список в словарь
    prompts_dict = {}
    for prompt in prompts_data:
        prompt_id = prompt['system_prompt_id']
        prompts_dict[prompt_id] = {
            'prompt': prompt['system_prompt'],
            'description': prompt.get('description', ''),
            'version': prompt.get('version', ''),
        }

    return prompts_dict

def load_ragchecker_metrics_for_version(version_path, version_name):
    """Загрузка метрик RAGChecker для одной версии"""
    metrics_data = []
    ragchecker_path = Path(version_path) / "RAGChecker_outputs"

    if not ragchecker_path.exists():
        print(f"Путь не существует: {ragchecker_path}")
        return pd.DataFrame()

    # Перебираем все модели
    for model_dir in ragchecker_path.iterdir():
        if not model_dir.is_dir():
            continue

        model_name = model_dir.name

        # Перебираем все промпты
        for prompt_dir in model_dir.iterdir():
            if not prompt_dir.is_dir():
                continue

            prompt_id = prompt_dir.name

            # Ищем JSON файл с метриками
            json_files = list(prompt_dir.glob("ragchecker_report_*.json"))

            if not json_files:
                continue

            json_file = json_files[0]

            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Извлекаем метрики
                overall = data.get('overall_metrics', {})
                retriever = data.get('retriever_metrics', {})
                generator = data.get('generator_metrics', {})

                metrics_data.append({
                    'version': version_name,
                    'model_name': model_name,
                    'prompt_id': prompt_id,
                    # Overall metrics
                    'precision': overall.get('precision', None),
                    'recall': overall.get('recall', None),
                    'f1': overall.get('f1', None),
                    # Retriever metrics
                    'claim_recall': retriever.get('claim_recall', None),
                    'context_precision': retriever.get('context_precision', None),
                    # Generator metrics
                    'context_utilization': generator.get('context_utilization', None),
                    'noise_sensitivity_relevant': generator.get('noise_sensitivity_in_relevant', None),
                    'noise_sensitivity_irrelevant': generator.get('noise_sensitivity_in_irrelevant', None),
                    'hallucination': generator.get('hallucination', None),
                    'self_knowledge': generator.get('self_knowledge', None),
                    'faithfulness': generator.get('faithfulness', None),
                })

            except Exception as e:
                print(f"Ошибка при загрузке {json_file}: {e}")

    return pd.DataFrame(metrics_data)

def compare_metrics(df_v1, df_v2):
    """Сравнение метрик между версиями"""

    # Объединяем данные
    df_v1['version'] = 'v1_english'
    df_v2['version'] = 'v2_russian'

    # Мержим по модели и промпту
    comparison = pd.merge(
        df_v1, df_v2,
        on=['model_name', 'prompt_id'],
        suffixes=('_v1', '_v2'),
        how='outer'
    )

    # Вычисляем разницу для ключевых метрик
    metrics_to_compare = ['precision', 'recall', 'f1', 'hallucination', 'faithfulness',
                         'context_utilization', 'claim_recall', 'context_precision']

    for metric in metrics_to_compare:
        col_v1 = f'{metric}_v1'
        col_v2 = f'{metric}_v2'
        col_diff = f'{metric}_diff'
        col_pct = f'{metric}_change_pct'

        if col_v1 in comparison.columns and col_v2 in comparison.columns:
            comparison[col_diff] = comparison[col_v2] - comparison[col_v1]
            comparison[col_pct] = (
                (comparison[col_v2] - comparison[col_v1]) /
                comparison[col_v1].replace(0, np.nan) * 100
            ).round(2)

    return comparison

def create_comparison_excel(version1_path, version2_path, output_file):
    """Создание Excel отчета со сравнением версий"""

    print("="*70)
    print("Сравнение двух версий RAGChecker")
    print("="*70)

    # Загружаем промпты
    print("\n1. Загрузка промптов...")
    prompts_v1 = load_prompts(version1_path)
    prompts_v2 = load_prompts(version2_path)

    print(f"   Version 1: {len(prompts_v1)} промптов")
    print(f"   Version 2: {len(prompts_v2)} промптов")

    # Создаем DataFrame с промптами
    prompts_comparison = []
    for prompt_id in sorted(set(list(prompts_v1.keys()) + list(prompts_v2.keys()))):
        prompts_comparison.append({
            'prompt_id': prompt_id,
            'v1_prompt': prompts_v1.get(prompt_id, {}).get('prompt', 'N/A'),
            'v1_description': prompts_v1.get(prompt_id, {}).get('description', 'N/A'),
            'v1_version': prompts_v1.get(prompt_id, {}).get('version', 'N/A'),
            'v2_prompt': prompts_v2.get(prompt_id, {}).get('prompt', 'N/A'),
            'v2_description': prompts_v2.get(prompt_id, {}).get('description', 'N/A'),
            'v2_version': prompts_v2.get(prompt_id, {}).get('version', 'N/A'),
        })

    df_prompts = pd.DataFrame(prompts_comparison)

    # Загружаем метрики
    print("\n2. Загрузка метрик Version 1...")
    df_v1 = load_ragchecker_metrics_for_version(version1_path, 'v1_english')
    print(f"   Загружено: {len(df_v1)} записей")

    print("\n3. Загрузка метрик Version 2...")
    df_v2 = load_ragchecker_metrics_for_version(version2_path, 'v2_russian')
    print(f"   Загружено: {len(df_v2)} записей")

    # Сравниваем метрики
    print("\n4. Сравнение метрик...")
    df_comparison = compare_metrics(df_v1, df_v2)

    # Создаем сводки
    print("\n5. Создание сводных таблиц...")

    # Сводка по моделям
    model_summary_v1 = df_v1.groupby('model_name').agg({
        'precision': 'mean', 'recall': 'mean', 'f1': 'mean',
        'hallucination': 'mean', 'faithfulness': 'mean'
    }).round(2).add_suffix('_v1')

    model_summary_v2 = df_v2.groupby('model_name').agg({
        'precision': 'mean', 'recall': 'mean', 'f1': 'mean',
        'hallucination': 'mean', 'faithfulness': 'mean'
    }).round(2).add_suffix('_v2')

    model_summary = pd.merge(
        model_summary_v1, model_summary_v2,
        left_index=True, right_index=True, how='outer'
    )

    # Добавляем разницу
    for metric in ['precision', 'recall', 'f1', 'hallucination', 'faithfulness']:
        model_summary[f'{metric}_diff'] = (
            model_summary[f'{metric}_v2'] - model_summary[f'{metric}_v1']
        ).round(2)

    # Сводка по промптам
    prompt_summary_v1 = df_v1.groupby('prompt_id').agg({
        'precision': 'mean', 'recall': 'mean', 'f1': 'mean',
        'hallucination': 'mean', 'faithfulness': 'mean'
    }).round(2).add_suffix('_v1')

    prompt_summary_v2 = df_v2.groupby('prompt_id').agg({
        'precision': 'mean', 'recall': 'mean', 'f1': 'mean',
        'hallucination': 'mean', 'faithfulness': 'mean'
    }).round(2).add_suffix('_v2')

    prompt_summary = pd.merge(
        prompt_summary_v1, prompt_summary_v2,
        left_index=True, right_index=True, how='outer'
    )

    # Добавляем разницу
    for metric in ['precision', 'recall', 'f1', 'hallucination', 'faithfulness']:
        prompt_summary[f'{metric}_diff'] = (
            prompt_summary[f'{metric}_v2'] - prompt_summary[f'{metric}_v1']
        ).round(2)

    # Матрицы сравнения F1
    pivot_f1_v1 = df_v1.pivot_table(
        values='f1', index='model_name', columns='prompt_id', aggfunc='mean'
    ).round(2)

    pivot_f1_v2 = df_v2.pivot_table(
        values='f1', index='model_name', columns='prompt_id', aggfunc='mean'
    ).round(2)

    pivot_f1_diff = (pivot_f1_v2 - pivot_f1_v1).round(2)

    # Создаем Excel файл
    print("\n6. Создание Excel файла...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # 1. Сравнение промптов
        df_prompts.to_excel(writer, sheet_name='Сравнение промптов', index=False)

        # 2. Детальное сравнение метрик
        df_comparison.to_excel(writer, sheet_name='Детальное сравнение', index=False)

        # 3. Сводка по моделям
        model_summary.to_excel(writer, sheet_name='Модели - сравнение')

        # 4. Сводка по промптам
        prompt_summary.to_excel(writer, sheet_name='Промпты - сравнение')

        # 5. F1 Version 1
        pivot_f1_v1.to_excel(writer, sheet_name='F1 - Version 1 (EN)')

        # 6. F1 Version 2
        pivot_f1_v2.to_excel(writer, sheet_name='F1 - Version 2 (RU)')

        # 7. F1 Difference
        pivot_f1_diff.to_excel(writer, sheet_name='F1 - Разница (V2-V1)')

        # 8. Только улучшения
        improvements = df_comparison[df_comparison['f1_diff'] > 0].sort_values('f1_diff', ascending=False)
        improvements[['model_name', 'prompt_id', 'f1_v1', 'f1_v2', 'f1_diff', 'f1_change_pct']].to_excel(
            writer, sheet_name='Улучшения', index=False
        )

        # 9. Только ухудшения
        degradations = df_comparison[df_comparison['f1_diff'] < 0].sort_values('f1_diff')
        degradations[['model_name', 'prompt_id', 'f1_v1', 'f1_v2', 'f1_diff', 'f1_change_pct']].to_excel(
            writer, sheet_name='Ухудшения', index=False
        )

        # 10. Статистика изменений
        changes_stats = pd.DataFrame({
            'Метрика': ['F1', 'Precision', 'Recall', 'Hallucination', 'Faithfulness'],
            'Среднее V1': [
                df_v1['f1'].mean(),
                df_v1['precision'].mean(),
                df_v1['recall'].mean(),
                df_v1['hallucination'].mean(),
                df_v1['faithfulness'].mean(),
            ],
            'Среднее V2': [
                df_v2['f1'].mean(),
                df_v2['precision'].mean(),
                df_v2['recall'].mean(),
                df_v2['hallucination'].mean(),
                df_v2['faithfulness'].mean(),
            ],
        })
        changes_stats['Изменение'] = (changes_stats['Среднее V2'] - changes_stats['Среднее V1']).round(2)
        changes_stats['Изменение %'] = (
            (changes_stats['Среднее V2'] - changes_stats['Среднее V1']) /
            changes_stats['Среднее V1'].replace(0, np.nan) * 100
        ).round(2)
        changes_stats.to_excel(writer, sheet_name='Общая статистика', index=False)

    # Применяем форматирование
    print("\n7. Применение условного форматирования...")
    wb = load_workbook(output_file)

    format_prompts_sheet(wb['Сравнение промптов'])
    format_comparison_sheet(wb['Детальное сравнение'])
    format_summary_with_diff(wb['Модели - сравнение'])
    format_summary_with_diff(wb['Промпты - сравнение'])
    format_matrix_sheet(wb['F1 - Version 1 (EN)'], 'percent')
    format_matrix_sheet(wb['F1 - Version 2 (RU)'], 'percent')
    format_diff_matrix_sheet(wb['F1 - Разница (V2-V1)'])
    format_improvements_sheet(wb['Улучшения'])
    format_degradations_sheet(wb['Ухудшения'])
    format_stats_sheet(wb['Общая статистика'])

    wb.save(output_file)

    print(f"\n✅ Отчет сохранен: {output_file}")

    # Выводим статистику
    print("\n" + "="*70)
    print("РЕЗУЛЬТАТЫ СРАВНЕНИЯ")
    print("="*70)

    print(f"\nОбщая статистика:")
    print(f"  F1:")
    print(f"    Version 1 (EN): {df_v1['f1'].mean():.2f}")
    print(f"    Version 2 (RU): {df_v2['f1'].mean():.2f}")
    print(f"    Изменение: {df_v2['f1'].mean() - df_v1['f1'].mean():.2f}")

    print(f"\n  Precision:")
    print(f"    Version 1 (EN): {df_v1['precision'].mean():.2f}")
    print(f"    Version 2 (RU): {df_v2['precision'].mean():.2f}")
    print(f"    Изменение: {df_v2['precision'].mean() - df_v1['precision'].mean():.2f}")

    print(f"\n  Recall:")
    print(f"    Version 1 (EN): {df_v1['recall'].mean():.2f}")
    print(f"    Version 2 (RU): {df_v2['recall'].mean():.2f}")
    print(f"    Изменение: {df_v2['recall'].mean() - df_v1['recall'].mean():.2f}")

    improvements_count = len(improvements)
    degradations_count = len(degradations)
    total = len(df_comparison)

    print(f"\nРаспределение изменений F1:")
    print(f"  Улучшений: {improvements_count} ({improvements_count/total*100:.1f}%)")
    print(f"  Ухудшений: {degradations_count} ({degradations_count/total*100:.1f}%)")
    print(f"  Без изменений: {total - improvements_count - degradations_count}")

    if not improvements.empty:
        best_improvement = improvements.iloc[0]
        print(f"\nНаибольшее улучшение:")
        print(f"  Модель: {best_improvement['model_name']}")
        print(f"  Промпт: {best_improvement['prompt_id']}")
        print(f"  F1: {best_improvement['f1_v1']:.2f} → {best_improvement['f1_v2']:.2f} (+{best_improvement['f1_diff']:.2f})")

    if not degradations.empty:
        worst_degradation = degradations.iloc[0]
        print(f"\nНаибольшее ухудшение:")
        print(f"  Модель: {worst_degradation['model_name']}")
        print(f"  Промпт: {worst_degradation['prompt_id']}")
        print(f"  F1: {worst_degradation['f1_v1']:.2f} → {worst_degradation['f1_v2']:.2f} ({worst_degradation['f1_diff']:.2f})")

    print("\n" + "="*70)

# Функции форматирования
def format_prompts_sheet(ws):
    """Форматирование листа с промптами"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'A2'

def format_comparison_sheet(ws):
    """Форматирование листа детального сравнения"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Найдем колонки с _diff
    header_row = [cell.value for cell in ws[1]]
    for idx, col_name in enumerate(header_row, 1):
        if col_name and '_diff' in str(col_name):
            col_letter = get_column_letter(idx)
            # Зеленый для положительных, красный для отрицательных
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=-50, start_color='FFC7CE',
                    mid_type='num', mid_value=0, mid_color='FFEB9C',
                    end_type='num', end_value=50, end_color='C6EFCE'
                )
            )

    ws.freeze_panes = 'A2'

def format_summary_with_diff(ws):
    """Форматирование сводных листов с разницей"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws['A']:
        if cell.row > 1:
            cell.fill = LIGHT_BLUE_FILL
            cell.font = Font(bold=True)

    header_row = [cell.value for cell in ws[1]]
    for idx, col_name in enumerate(header_row, 1):
        if col_name and '_diff' in str(col_name):
            col_letter = get_column_letter(idx)
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=-30, start_color='FFC7CE',
                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                    end_type='num', end_value=30, end_color='C6EFCE'
                )
            )

    ws.freeze_panes = 'B2'

def format_matrix_sheet(ws, metric_type='percent'):
    """Форматирование матричных листов"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws['A']:
        if cell.row > 1:
            cell.fill = LIGHT_BLUE_FILL
            cell.font = Font(bold=True)

    if ws.max_row > 1 and ws.max_column > 1:
        start_col = get_column_letter(2)
        end_col = get_column_letter(ws.max_column)

        ws.conditional_formatting.add(
            f'{start_col}2:{end_col}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FFC7CE',
                mid_type='num', mid_value=50, mid_color='FFEB9C',
                end_type='num', end_value=100, end_color='C6EFCE'
            )
        )

    ws.freeze_panes = 'B2'

def format_diff_matrix_sheet(ws):
    """Форматирование матрицы разниц"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws['A']:
        if cell.row > 1:
            cell.fill = LIGHT_BLUE_FILL
            cell.font = Font(bold=True)

    if ws.max_row > 1 and ws.max_column > 1:
        start_col = get_column_letter(2)
        end_col = get_column_letter(ws.max_column)

        # Разница: отрицательные красные, положительные зеленые
        ws.conditional_formatting.add(
            f'{start_col}2:{end_col}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=-50, start_color='FF0000',
                mid_type='num', mid_value=0, mid_color='FFFFFF',
                end_type='num', end_value=50, end_color='00B050'
            )
        )

    ws.freeze_panes = 'B2'

def format_improvements_sheet(ws):
    """Форматирование листа улучшений"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Подсветка всех улучшений зеленым
    if ws.max_row > 1:
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = LIGHT_GREEN_FILL

    ws.freeze_panes = 'A2'

def format_degradations_sheet(ws):
    """Форматирование листа ухудшений"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Подсветка всех ухудшений красным
    if ws.max_row > 1:
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = RED_FILL

    ws.freeze_panes = 'A2'

def format_stats_sheet(ws):
    """Форматирование листа статистики"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Найдем колонку "Изменение"
    header_row = [cell.value for cell in ws[1]]
    if 'Изменение' in header_row:
        col_idx = header_row.index('Изменение') + 1
        col_letter = get_column_letter(col_idx)

        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            ColorScaleRule(
                start_type='num', start_value=-30, start_color='FFC7CE',
                mid_type='num', mid_value=0, mid_color='FFFFFF',
                end_type='num', end_value=30, end_color='C6EFCE'
            )
        )

    ws.freeze_panes = 'A2'

if __name__ == '__main__':
    version1_path = ' version_1'  # Папка с пробелом в начале
    version2_path = 'version_2'
    output_file = 'versions_comparison.xlsx'

    create_comparison_excel(version1_path, version2_path, output_file)
