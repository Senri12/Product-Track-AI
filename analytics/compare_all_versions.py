#!/usr/bin/env python3
"""
Скрипт для сравнения ТРЕХ версий данных RAGChecker
Version 1: английские промпты
Version 2: русские промпты (первая итерация)
Version 3: русские промпты (вторая итерация)

С добавлением описаний метрик и сравнением с бенчмарком из статьи
"""

import pandas as pd
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.comments import Comment
import numpy as np

# Цвета
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
LIGHT_GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

# Описания метрик
METRICS_DESCRIPTIONS = {
    'precision': 'Точность: доля правильных утверждений в ответе модели. Высокое значение = мало лишней информации',
    'recall': 'Полнота: доля эталонных утверждений, покрытых моделью. Высокое значение = полный ответ',
    'f1': 'F1-мера: баланс точности и полноты. ГЛАВНАЯ МЕТРИКА КАЧЕСТВА. >60% - отлично, 40-60% - хорошо',
    'claim_recall': 'Claim Recall: доля эталонных утверждений в извлеченных чанках. Оценка качества retriever',
    'context_precision': 'Context Precision: доля релевантных чанков среди извлеченных. Высокое значение = мало шума',
    'context_utilization': 'Context Utilization: насколько эффективно модель использует предоставленный контекст',
    'noise_sensitivity_relevant': 'Relevant Noise Sensitivity: чувствительность к шуму в релевантном контексте. Низкое = хорошо',
    'noise_sensitivity_irrelevant': 'Irrelevant Noise Sensitivity: чувствительность к нерелевантному контексту. Низкое = хорошо',
    'hallucination': 'Hallucination: доля "изобретенных" фактов. КРИТИЧНО! <10% - отлично, >50% - неприемлемо',
    'self_knowledge': 'Self-knowledge: использование собственных знаний модели. Может быть + или - в зависимости от задачи',
    'faithfulness': 'Faithfulness: точность следования контексту. >80% - отлично. Связано с hallucination'
}

# Бенчмарк из статьи (ClapNQ, лучшие системы)
BENCHMARK_DATA = {
    'E5-Mistral_GPT-4': {
        'precision': 59.7, 'recall': 51.1, 'f1': 47.9,
        'claim_recall': 81.5, 'context_precision': 43.6,
        'context_utilization': 59.9,
        'noise_sensitivity_relevant': 31.1,
        'noise_sensitivity_irrelevant': 3.8,
        'hallucination': 5.4,
        'self_knowledge': 2.3,
        'faithfulness': 92.3
    },
    'BM25_GPT-4': {
        'precision': 56.9, 'recall': 50.0, 'f1': 46.7,
        'claim_recall': 81.1, 'context_precision': 41.3,
        'context_utilization': 56.4,
        'noise_sensitivity_relevant': 29.4,
        'noise_sensitivity_irrelevant': 5.9,
        'hallucination': 7.5,
        'self_knowledge': 2.2,
        'faithfulness': 90.3
    }
}

def load_prompts(version_path):
    """Загрузка промптов из JSON файла"""
    prompts_file = Path(version_path) / "system_prompts.json"

    if not prompts_file.exists():
        print(f"Файл не найден: {prompts_file}")
        return {}

    with open(prompts_file, 'r', encoding='utf-8') as f:
        prompts_data = json.load(f)

    prompts_dict = {}
    for prompt in prompts_data:
        prompt_id = prompt['system_prompt_id']
        prompts_dict[prompt_id] = {
            'prompt': prompt['system_prompt'],
            'description': prompt.get('description', ''),
            'version': prompt.get('version', ''),
        }

    return prompts_dict

def load_ragchecker_metrics_for_version(version_path, version_name):
    """Загрузка метрик RAGChecker для одной версии"""
    metrics_data = []
    ragchecker_path = Path(version_path) / "RAGChecker_outputs"

    if not ragchecker_path.exists():
        print(f"Путь не существует: {ragchecker_path}")
        return pd.DataFrame()

    for model_dir in ragchecker_path.iterdir():
        if not model_dir.is_dir():
            continue

        model_name = model_dir.name

        for prompt_dir in model_dir.iterdir():
            if not prompt_dir.is_dir():
                continue

            prompt_id = prompt_dir.name
            json_files = list(prompt_dir.glob("ragchecker_report_*.json"))

            if not json_files:
                continue

            json_file = json_files[0]

            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                overall = data.get('overall_metrics', {})
                retriever = data.get('retriever_metrics', {})
                generator = data.get('generator_metrics', {})

                metrics_data.append({
                    'version': version_name,
                    'model_name': model_name,
                    'prompt_id': prompt_id,
                    'precision': overall.get('precision', None),
                    'recall': overall.get('recall', None),
                    'f1': overall.get('f1', None),
                    'claim_recall': retriever.get('claim_recall', None),
                    'context_precision': retriever.get('context_precision', None),
                    'context_utilization': generator.get('context_utilization', None),
                    'noise_sensitivity_relevant': generator.get('noise_sensitivity_in_relevant', None),
                    'noise_sensitivity_irrelevant': generator.get('noise_sensitivity_in_irrelevant', None),
                    'hallucination': generator.get('hallucination', None),
                    'self_knowledge': generator.get('self_knowledge', None),
                    'faithfulness': generator.get('faithfulness', None),
                })

            except Exception as e:
                print(f"Ошибка при загрузке {json_file}: {e}")

    return pd.DataFrame(metrics_data)

def create_comparison_excel(version_paths, output_file):
    """Создание Excel отчета со сравнением всех версий"""

    print("="*70)
    print("Сравнение ТРЕХ версий RAGChecker")
    print("="*70)

    # Загружаем данные для всех версий
    all_versions_data = {}
    all_prompts = {}

    for ver_name, ver_path in version_paths.items():
        print(f"\n{ver_name}: Загрузка промптов...")
        prompts = load_prompts(ver_path)
        all_prompts[ver_name] = prompts
        print(f"  {len(prompts)} промптов")

        print(f"{ver_name}: Загрузка метрик...")
        df = load_ragchecker_metrics_for_version(ver_path, ver_name)
        all_versions_data[ver_name] = df
        print(f"  {len(df)} записей")

    # Создаем сравнительные таблицы
    print("\nСоздание сводных таблиц...")

    # Объединяем все версии
    df_all = pd.concat(all_versions_data.values(), ignore_index=True)

    # Создаем Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # 1. Описание метрик
        metrics_desc_df = pd.DataFrame([
            {'Метрика': k, 'Описание': v, 'Категория': get_metric_category(k)}
            for k, v in METRICS_DESCRIPTIONS.items()
        ])
        metrics_desc_df.to_excel(writer, sheet_name='📖 Описание метрик', index=False)

        # 2. Сравнение с бенчмарком
        benchmark_df = pd.DataFrame(BENCHMARK_DATA).T
        benchmark_df['source'] = 'RAGChecker Paper (ClapNQ)'

        # Добавляем наши результаты
        our_results = {}
        for ver_name, df in all_versions_data.items():
            our_results[ver_name] = df.mean(numeric_only=True).to_dict()
            our_results[ver_name]['source'] = f'Our results ({ver_name})'

        our_results_df = pd.DataFrame(our_results).T
        comparison_with_benchmark = pd.concat([benchmark_df, our_results_df])
        comparison_with_benchmark.to_excel(writer, sheet_name='📊 vs Бенчмарк')

        # 3. Общая статистика по версиям
        version_summary = df_all.groupby('version').agg({
            'precision': ['mean', 'std'],
            'recall': ['mean', 'std'],
            'f1': ['mean', 'std'],
            'hallucination': ['mean', 'std'],
            'faithfulness': ['mean', 'std']
        }).round(2)
        version_summary.to_excel(writer, sheet_name='📈 Сводка по версиям')

        # 4. Детальное сравнение всех версий
        df_all.to_excel(writer, sheet_name='📋 Все данные', index=False)

        # 5. Сравнение V1 vs V2
        df_v1 = all_versions_data.get('v1_english', pd.DataFrame())
        df_v2 = all_versions_data.get('v2_russian', pd.DataFrame())
        if not df_v1.empty and not df_v2.empty:
            comparison_v1_v2 = compare_two_versions(df_v1, df_v2, 'v1', 'v2')
            comparison_v1_v2.to_excel(writer, sheet_name='V1 vs V2', index=False)

        # 6. Сравнение V2 vs V3
        df_v3 = all_versions_data.get('v3_russian_v2', pd.DataFrame())
        if not df_v2.empty and not df_v3.empty:
            comparison_v2_v3 = compare_two_versions(df_v2, df_v3, 'v2', 'v3')
            comparison_v2_v3.to_excel(writer, sheet_name='V2 vs V3', index=False)

        # 7. Сравнение V1 vs V3
        if not df_v1.empty and not df_v3.empty:
            comparison_v1_v3 = compare_two_versions(df_v1, df_v3, 'v1', 'v3')
            comparison_v1_v3.to_excel(writer, sheet_name='V1 vs V3', index=False)

        # 8. Матрица F1 по версиям
        pivot_f1 = df_all.pivot_table(
            values='f1',
            index=['model_name', 'prompt_id'],
            columns='version',
            aggfunc='mean'
        ).round(2)
        pivot_f1.to_excel(writer, sheet_name='F1 по версиям')

        # 9. Лучшие улучшения
        if 'v1_english' in all_versions_data and 'v3_russian_v2' in all_versions_data:
            improvements = find_improvements(df_v1, df_v3)
            improvements.to_excel(writer, sheet_name='🟢 Улучшения V1→V3', index=False)

        # 10. Сравнение промптов
        prompts_comparison = create_prompts_comparison(all_prompts)
        prompts_comparison.to_excel(writer, sheet_name='📝 Промпты', index=False)

    # Применяем форматирование
    print("\nПрименение форматирования...")
    wb = load_workbook(output_file)

    format_metrics_description_sheet(wb['📖 Описание метрик'])
    format_benchmark_sheet(wb['📊 vs Бенчмарк'])
    format_version_summary_sheet(wb['📈 Сводка по версиям'])

    # Добавляем комментарии к заголовкам
    add_metric_comments(wb['📋 Все данные'])

    wb.save(output_file)

    print(f"\n✅ Отчет сохранен: {output_file}")
    print_summary_stats(all_versions_data)

def get_metric_category(metric_name):
    """Определить категорию метрики"""
    if metric_name in ['precision', 'recall', 'f1']:
        return 'Overall'
    elif metric_name in ['claim_recall', 'context_precision']:
        return 'Retriever'
    else:
        return 'Generator'

def compare_two_versions(df1, df2, name1, name2):
    """Сравнение двух версий"""
    merged = pd.merge(
        df1, df2,
        on=['model_name', 'prompt_id'],
        suffixes=(f'_{name1}', f'_{name2}'),
        how='outer'
    )

    metrics = ['precision', 'recall', 'f1', 'hallucination', 'faithfulness']
    for metric in metrics:
        col1 = f'{metric}_{name1}'
        col2 = f'{metric}_{name2}'
        if col1 in merged.columns and col2 in merged.columns:
            merged[f'{metric}_diff'] = merged[col2] - merged[col1]
            merged[f'{metric}_pct'] = ((merged[col2] - merged[col1]) / merged[col1].replace(0, np.nan) * 100).round(2)

    return merged

def find_improvements(df1, df3):
    """Найти улучшения между V1 и V3"""
    comparison = compare_two_versions(df1, df3, 'v1', 'v3')
    improvements = comparison[comparison['f1_diff'] > 0].sort_values('f1_diff', ascending=False)
    return improvements[['model_name', 'prompt_id', 'f1_v1', 'f1_v3', 'f1_diff', 'f1_pct']]

def create_prompts_comparison(all_prompts):
    """Создать таблицу сравнения промптов"""
    rows = []
    prompt_ids = set()
    for prompts in all_prompts.values():
        prompt_ids.update(prompts.keys())

    for prompt_id in sorted(prompt_ids):
        row = {'prompt_id': prompt_id}
        for ver_name, prompts in all_prompts.items():
            if prompt_id in prompts:
                row[f'{ver_name}_prompt'] = prompts[prompt_id]['prompt']
                row[f'{ver_name}_desc'] = prompts[prompt_id]['description']
        rows.append(row)

    return pd.DataFrame(rows)

def format_metrics_description_sheet(ws):
    """Форматирование листа описания метрик"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Авто-ширина
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 15

    ws.freeze_panes = 'A2'

def format_benchmark_sheet(ws):
    """Форматирование листа с бенчмарком"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Подсветка строк бенчмарка
    for row in range(2, ws.max_row + 1):
        source_cell = ws.cell(row, ws.max_column)
        if 'Paper' in str(source_cell.value):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = LIGHT_BLUE_FILL
                ws.cell(row, col).font = Font(bold=True)

    ws.freeze_panes = 'B2'

def format_version_summary_sheet(ws):
    """Форматирование сводки по версиям"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = 'B2'

def add_metric_comments(ws):
    """Добавить комментарии к заголовкам метрик"""
    header_row = [cell.value for cell in ws[1]]

    for idx, col_name in enumerate(header_row, 1):
        if col_name in METRICS_DESCRIPTIONS:
            cell = ws.cell(1, idx)
            comment = Comment(METRICS_DESCRIPTIONS[col_name], 'RAGChecker')
            cell.comment = comment
            cell.fill = YELLOW_FILL  # Желтая заливка для колонок с комментариями

def print_summary_stats(all_versions_data):
    """Вывести сводную статистику"""
    print("\n" + "="*70)
    print("СВОДНАЯ СТАТИСТИКА ПО ВЕРСИЯМ")
    print("="*70)

    for ver_name, df in all_versions_data.items():
        if df.empty:
            continue
        print(f"\n{ver_name}:")
        print(f"  F1:           {df['f1'].mean():.2f}%")
        print(f"  Precision:    {df['precision'].mean():.2f}%")
        print(f"  Recall:       {df['recall'].mean():.2f}%")
        print(f"  Hallucination: {df['hallucination'].mean():.2f}%")
        print(f"  Faithfulness: {df['faithfulness'].mean():.2f}%")

    print("\n" + "="*70)

if __name__ == '__main__':
    version_paths = {
        'v1_english': ' version_1',  # Пробел в начале!
        'v2_russian': 'version_2',
        'v3_russian_v2': 'version_3'
    }

    output_file = 'comparison_all_versions.xlsx'

    create_comparison_excel(version_paths, output_file)
