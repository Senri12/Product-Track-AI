#!/usr/bin/env python3
"""
Скрипт для анализа метрик RAGChecker и создания Excel отчета
с сопоставлением метрик по промптам и моделям
"""

import pandas as pd
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import numpy as np

# Цвета для условного форматирования
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

def load_ragchecker_metrics(base_path):
    """
    Загрузка всех метрик RAGChecker из папки RAGChecker_outputs
    """
    metrics_data = []
    ragchecker_path = Path(base_path) / "RAGChecker_outputs"

    if not ragchecker_path.exists():
        print(f"Путь не существует: {ragchecker_path}")
        return pd.DataFrame()

    # Перебираем все модели
    for model_dir in ragchecker_path.iterdir():
        if not model_dir.is_dir():
            continue

        model_name = model_dir.name

        # Перебираем все промпты
        for prompt_dir in model_dir.iterdir():
            if not prompt_dir.is_dir():
                continue

            prompt_id = prompt_dir.name

            # Ищем JSON файл с метриками
            json_files = list(prompt_dir.glob("ragchecker_report_*.json"))

            if not json_files:
                print(f"Не найден JSON для {model_name}/{prompt_id}")
                continue

            json_file = json_files[0]

            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Извлекаем метрики
                overall = data.get('overall_metrics', {})
                retriever = data.get('retriever_metrics', {})
                generator = data.get('generator_metrics', {})

                metrics_data.append({
                    'model_name': model_name,
                    'prompt_id': prompt_id,
                    # Overall metrics
                    'precision': overall.get('precision', None),
                    'recall': overall.get('recall', None),
                    'f1': overall.get('f1', None),
                    # Retriever metrics
                    'claim_recall': retriever.get('claim_recall', None),
                    'context_precision': retriever.get('context_precision', None),
                    # Generator metrics
                    'context_utilization': generator.get('context_utilization', None),
                    'noise_sensitivity_relevant': generator.get('noise_sensitivity_in_relevant', None),
                    'noise_sensitivity_irrelevant': generator.get('noise_sensitivity_in_irrelevant', None),
                    'hallucination': generator.get('hallucination', None),
                    'self_knowledge': generator.get('self_knowledge', None),
                    'faithfulness': generator.get('faithfulness', None),
                })

                print(f"✓ Загружено: {model_name}/{prompt_id}")

            except Exception as e:
                print(f"✗ Ошибка при загрузке {json_file}: {e}")

    return pd.DataFrame(metrics_data)

def load_model_outputs_stats(base_path):
    """
    Загрузка статистики из model_outputs (входные данные для RAGChecker)
    """
    stats_data = []
    model_outputs_path = Path(base_path) / "model_outputs"

    if not model_outputs_path.exists():
        print(f"Путь не существует: {model_outputs_path}")
        return pd.DataFrame()

    # Перебираем все модели
    for model_dir in model_outputs_path.iterdir():
        if not model_dir.is_dir():
            continue

        model_name = model_dir.name

        # Перебираем все промпты
        for prompt_dir in model_dir.iterdir():
            if not prompt_dir.is_dir():
                continue

            prompt_id = prompt_dir.name

            # Считаем количество диалогов
            dialog_files = list(prompt_dir.glob("dialog*.json"))
            num_dialogs = len(dialog_files)

            # Анализируем диалоги
            total_questions = 0
            total_answers_length = 0
            has_russian = 0
            has_english = 0

            for dialog_file in dialog_files[:10]:  # Берем первые 10 для статистики
                try:
                    with open(dialog_file, 'r', encoding='utf-8') as f:
                        dialog_data = json.load(f)

                    # Считаем вопросы
                    if 'messages' in dialog_data:
                        for msg in dialog_data['messages']:
                            if msg.get('role') == 'user':
                                total_questions += 1
                            elif msg.get('role') == 'assistant':
                                content = msg.get('content', '')
                                total_answers_length += len(content)

                                # Определяем язык
                                if any(ord(c) >= 0x0400 and ord(c) <= 0x04FF for c in content):
                                    has_russian += 1
                                if any(c.isalpha() and ord(c) < 128 for c in content[:100]):
                                    has_english += 1

                except Exception as e:
                    pass

            stats_data.append({
                'model_name': model_name,
                'prompt_id': prompt_id,
                'num_dialogs': num_dialogs,
                'avg_questions_per_dialog': total_questions / max(len(dialog_files[:10]), 1),
                'avg_answer_length': total_answers_length / max(has_russian + has_english, 1),
                'russian_answers': has_russian,
                'english_answers': has_english,
            })

            print(f"✓ Статистика: {model_name}/{prompt_id} - {num_dialogs} диалогов")

    return pd.DataFrame(stats_data)

def create_excel_report(metrics_df, stats_df, output_file):
    """
    Создание Excel отчета с метриками и условным форматированием
    """

    print("\nСоздание Excel отчета...")

    # Объединяем метрики и статистику
    if not stats_df.empty:
        full_df = pd.merge(metrics_df, stats_df, on=['model_name', 'prompt_id'], how='left')
    else:
        full_df = metrics_df

    # Создаем Excel файл
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # 1. Лист со всеми метриками
        full_df.to_excel(writer, sheet_name='Все метрики', index=False)

        # 2. Сводка по моделям
        model_summary = metrics_df.groupby('model_name').agg({
            'precision': ['mean', 'std', 'min', 'max'],
            'recall': ['mean', 'std', 'min', 'max'],
            'f1': ['mean', 'std', 'min', 'max'],
            'hallucination': ['mean', 'std'],
            'faithfulness': ['mean', 'std'],
        }).round(2)
        model_summary.columns = ['_'.join(col).strip() for col in model_summary.columns.values]
        model_summary.to_excel(writer, sheet_name='Сводка по моделям')

        # 3. Сводка по промптам
        prompt_summary = metrics_df.groupby('prompt_id').agg({
            'precision': ['mean', 'std', 'min', 'max'],
            'recall': ['mean', 'std', 'min', 'max'],
            'f1': ['mean', 'std', 'min', 'max'],
            'hallucination': ['mean', 'std'],
            'faithfulness': ['mean', 'std'],
        }).round(2)
        prompt_summary.columns = ['_'.join(col).strip() for col in prompt_summary.columns.values]
        prompt_summary.to_excel(writer, sheet_name='Сводка по промптам')

        # 4. Матрица Precision (модели × промпты)
        pivot_precision = metrics_df.pivot_table(
            values='precision',
            index='model_name',
            columns='prompt_id',
            aggfunc='mean'
        ).round(2)
        pivot_precision.to_excel(writer, sheet_name='Precision (модель×промпт)')

        # 5. Матрица Recall
        pivot_recall = metrics_df.pivot_table(
            values='recall',
            index='model_name',
            columns='prompt_id',
            aggfunc='mean'
        ).round(2)
        pivot_recall.to_excel(writer, sheet_name='Recall (модель×промпт)')

        # 6. Матрица F1
        pivot_f1 = metrics_df.pivot_table(
            values='f1',
            index='model_name',
            columns='prompt_id',
            aggfunc='mean'
        ).round(2)
        pivot_f1.to_excel(writer, sheet_name='F1 (модель×промпт)')

        # 7. Матрица Hallucination
        pivot_hallucination = metrics_df.pivot_table(
            values='hallucination',
            index='model_name',
            columns='prompt_id',
            aggfunc='mean'
        ).round(2)
        pivot_hallucination.to_excel(writer, sheet_name='Hallucination (модель×промпт)')

        # 8. Матрица Faithfulness
        pivot_faithfulness = metrics_df.pivot_table(
            values='faithfulness',
            index='model_name',
            columns='prompt_id',
            aggfunc='mean'
        ).round(2)
        pivot_faithfulness.to_excel(writer, sheet_name='Faithfulness (модель×промпт)')

        # 9. Детальные метрики generator
        generator_metrics = metrics_df[['model_name', 'prompt_id', 'context_utilization',
                                        'noise_sensitivity_relevant', 'noise_sensitivity_irrelevant',
                                        'hallucination', 'self_knowledge', 'faithfulness']]
        generator_metrics.to_excel(writer, sheet_name='Generator метрики', index=False)

        # 10. Детальные метрики retriever
        retriever_metrics = metrics_df[['model_name', 'prompt_id', 'claim_recall', 'context_precision']]
        retriever_metrics.to_excel(writer, sheet_name='Retriever метрики', index=False)

    print("Применение условного форматирования...")

    # Открываем файл для форматирования
    wb = load_workbook(output_file)

    # Форматируем каждый лист
    format_all_metrics_sheet(wb['Все метрики'])
    format_summary_sheet(wb['Сводка по моделям'])
    format_summary_sheet(wb['Сводка по промптам'])

    # Форматируем матрицы
    format_matrix_sheet(wb['Precision (модель×промпт)'], metric_type='percent')
    format_matrix_sheet(wb['Recall (модель×промпт)'], metric_type='percent')
    format_matrix_sheet(wb['F1 (модель×промпт)'], metric_type='percent')
    format_matrix_sheet(wb['Hallucination (модель×промпт)'], metric_type='percent_inverse')
    format_matrix_sheet(wb['Faithfulness (модель×промпт)'], metric_type='percent')

    format_all_metrics_sheet(wb['Generator метрики'])
    format_all_metrics_sheet(wb['Retriever метрики'])

    wb.save(output_file)
    print(f"\n✓ Отчет сохранен: {output_file}")

def format_all_metrics_sheet(ws):
    """Форматирование листа с метриками"""

    # Заголовки
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Применяем цветовую шкалу к метрикам
    if ws.max_row > 1:
        header_row = [cell.value for cell in ws[1]]

        # Метрики, где больше = лучше (зеленые для высоких значений)
        good_metrics = ['precision', 'recall', 'f1', 'faithfulness', 'context_utilization',
                       'claim_recall', 'context_precision']

        # Метрики, где меньше = лучше (зеленые для низких значений)
        bad_metrics = ['hallucination', 'noise_sensitivity_relevant', 'noise_sensitivity_irrelevant']

        for idx, col_name in enumerate(header_row, 1):
            col_letter = get_column_letter(idx)

            if col_name in good_metrics:
                # Зеленый для высоких значений
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{ws.max_row}',
                    ColorScaleRule(
                        start_type='num', start_value=0, start_color='FFC7CE',
                        mid_type='num', mid_value=50, mid_color='FFEB9C',
                        end_type='num', end_value=100, end_color='C6EFCE'
                    )
                )
            elif col_name in bad_metrics:
                # Зеленый для низких значений (инвертированная шкала)
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{ws.max_row}',
                    ColorScaleRule(
                        start_type='num', start_value=0, start_color='C6EFCE',
                        mid_type='num', mid_value=50, mid_color='FFEB9C',
                        end_type='num', end_value=100, end_color='FFC7CE'
                    )
                )

    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'A2'

def format_summary_sheet(ws):
    """Форматирование сводных листов"""

    # Заголовки
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Первая колонка (названия)
    for cell in ws['A']:
        if cell.row > 1:
            cell.fill = LIGHT_BLUE_FILL
            cell.font = Font(bold=True)

    # Авто-ширина
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'B2'

def format_matrix_sheet(ws, metric_type='percent'):
    """
    Форматирование матричных листов
    metric_type: 'percent' (выше=лучше) или 'percent_inverse' (ниже=лучше)
    """

    # Заголовки
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Первая колонка (названия моделей)
    for cell in ws['A']:
        if cell.row > 1:
            cell.fill = LIGHT_BLUE_FILL
            cell.font = Font(bold=True)

    # Применить цветовую шкалу ко всем данным
    if ws.max_row > 1 and ws.max_column > 1:
        start_col = get_column_letter(2)
        end_col = get_column_letter(ws.max_column)

        if metric_type == 'percent_inverse':
            # Для hallucination: зеленый для низких, красный для высоких
            ws.conditional_formatting.add(
                f'{start_col}2:{end_col}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='C6EFCE',
                    mid_type='num', mid_value=50, mid_color='FFEB9C',
                    end_type='num', end_value=100, end_color='FFC7CE'
                )
            )
        else:
            # Для precision, recall, f1, faithfulness: зеленый для высоких
            ws.conditional_formatting.add(
                f'{start_col}2:{end_col}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='FFC7CE',
                    mid_type='num', mid_value=50, mid_color='FFEB9C',
                    end_type='num', end_value=100, end_color='C6EFCE'
                )
            )

    # Авто-ширина
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'B2'

if __name__ == '__main__':
    base_path = 'RAGChecker_outputs'
    output_file = 'ragchecker_analysis.xlsx'

    print("="*60)
    print("Анализ метрик RAGChecker")
    print("="*60)

    # Загружаем метрики RAGChecker
    print("\n1. Загрузка метрик RAGChecker...")
    metrics_df = load_ragchecker_metrics(base_path)

    if metrics_df.empty:
        print("✗ Не удалось загрузить метрики!")
        exit(1)

    print(f"\n✓ Загружено {len(metrics_df)} записей метрик")
    print(f"  Моделей: {metrics_df['model_name'].nunique()}")
    print(f"  Промптов: {metrics_df['prompt_id'].nunique()}")

    # Загружаем статистику входных данных
    print("\n2. Загрузка статистики входных данных...")
    stats_df = load_model_outputs_stats(base_path)

    if not stats_df.empty:
        print(f"✓ Загружено {len(stats_df)} записей статистики")

    # Создаем Excel отчет
    print("\n3. Создание Excel отчета...")
    create_excel_report(metrics_df, stats_df, output_file)

    print("\n" + "="*60)
    print("Готово!")
    print("="*60)

    # Выводим краткую статистику
    print("\nКраткая статистика метрик:")
    print("\nСредние значения по всем моделям и промптам:")
    print(f"  Precision: {metrics_df['precision'].mean():.2f}")
    print(f"  Recall: {metrics_df['recall'].mean():.2f}")
    print(f"  F1: {metrics_df['f1'].mean():.2f}")
    print(f"  Hallucination: {metrics_df['hallucination'].mean():.2f}")
    print(f"  Faithfulness: {metrics_df['faithfulness'].mean():.2f}")

    print("\nЛучшая модель по F1:")
    best_model = metrics_df.groupby('model_name')['f1'].mean().sort_values(ascending=False)
    print(f"  {best_model.index[0]}: {best_model.values[0]:.2f}")

    print("\nЛучший промпт по F1:")
    best_prompt = metrics_df.groupby('prompt_id')['f1'].mean().sort_values(ascending=False)
    print(f"  {best_prompt.index[0]}: {best_prompt.values[0]:.2f}")
