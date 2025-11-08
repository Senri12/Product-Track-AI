#!/usr/bin/env python3
"""
Скрипт для сравнения результатов экспериментов на разных лекциях
Автоматически определяет названия лекций из checking_inputs.json
"""

import pandas as pd
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
import numpy as np

# Цвета для условного форматирования
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')


def get_lecture_filenames(base_path):
    """
    Получение списка file_name из checking_inputs.json
    """
    checking_inputs_path = Path(base_path) / "checking_inputs.json"

    if not checking_inputs_path.exists():
        return []

    try:
        with open(checking_inputs_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return [item['file_name'] for item in data.get('data', [])]
    except Exception as e:
        print(f"  ⚠️  Ошибка при чтении {checking_inputs_path}: {e}")
        return []


def load_ragchecker_metrics(base_path, source_folder):
    """
    Загрузка метрик RAGChecker из папки

    Args:
        base_path: путь к папке версии (v1/v2/v3)
        source_folder: исходная папка ("analytics" или "2-file")
    """
    metrics_data = []

    # Получаем названия файлов лекций из checking_inputs.json
    lecture_filenames = get_lecture_filenames(base_path)

    if not lecture_filenames:
        print(f"  ⚠️  Не найдены file_name в checking_inputs.json для {base_path}")
        lecture_names = ["Unknown"]
    else:
        # Используем file_name как название лекции (может быть несколько)
        lecture_names = lecture_filenames

    # Определяем путь к RAGChecker_outputs в зависимости от источника
    if source_folder == "2-file":
        ragchecker_path = Path(base_path) / "RAGChecker_OUTPUTS"
    else:
        ragchecker_path = Path(base_path) / "RAGChecker_outputs"

    if not ragchecker_path.exists():
        print(f"  ⚠️  Путь не существует: {ragchecker_path}")
        return pd.DataFrame()

    # Перебираем все модели
    for model_dir in ragchecker_path.iterdir():
        if not model_dir.is_dir():
            continue

        model_name = model_dir.name

        # Перебираем все промпты
        for prompt_dir in model_dir.iterdir():
            if not prompt_dir.is_dir():
                continue

            prompt_id = prompt_dir.name

            # Ищем JSON файл с метриками
            json_files = list(prompt_dir.glob("ragchecker_report_*.json"))

            if not json_files:
                continue

            json_file = json_files[0]

            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # Извлекаем метрики
                overall = data.get('overall_metrics', {})
                retriever = data.get('retriever_metrics', {})
                generator = data.get('generator_metrics', {})

                # Создаем запись для каждого файла лекции
                for lecture_name in lecture_names:
                    metrics_data.append({
                        'lecture_file': lecture_name,
                        'source_folder': source_folder,
                        'version': base_path.name,
                        'model_name': model_name,
                        'prompt_id': prompt_id,
                        # Overall metrics
                        'precision': overall.get('precision', None),
                        'recall': overall.get('recall', None),
                        'f1': overall.get('f1', None),
                        # Retriever metrics
                        'claim_recall': retriever.get('claim_recall', None),
                        'context_precision': retriever.get('context_precision', None),
                        # Generator metrics
                        'context_utilization': generator.get('context_utilization', None),
                        'hallucination': generator.get('hallucination', None),
                        'faithfulness': generator.get('faithfulness', None),
                    })

            except Exception as e:
                print(f"  ✗ Ошибка при загрузке {json_file}: {e}")

    return pd.DataFrame(metrics_data)


def get_column_letter(col_idx):
    """Преобразование номера колонки в букву"""
    from openpyxl.utils import get_column_letter as gcl
    return gcl(col_idx)


def apply_conditional_formatting(worksheet, start_row, start_col, end_row, end_col, metric_type='higher_is_better'):
    """
    Применение условного форматирования к диапазону ячеек
    """
    if end_row <= start_row or end_col <= start_col:
        return

    if metric_type == 'higher_is_better':
        color_scale = ColorScaleRule(
            start_type='min', start_color='FFC7CE',
            mid_type='percentile', mid_value=50, mid_color='FFEB9C',
            end_type='max', end_color='C6EFCE'
        )
    else:
        color_scale = ColorScaleRule(
            start_type='min', start_color='C6EFCE',
            mid_type='percentile', mid_value=50, mid_color='FFEB9C',
            end_type='max', end_color='FFC7CE'
        )

    range_string = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    worksheet.conditional_formatting.add(range_string, color_scale)


def style_worksheet(worksheet):
    """Применение стилей к заголовкам"""
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def main():
    print("=" * 80)
    print("СРАВНИТЕЛЬНЫЙ АНАЛИЗ ЭКСПЕРИМЕНТОВ НА РАЗНЫХ ЛЕКЦИЯХ")
    print("=" * 80)

    # Определяем пути
    analytics_path = Path(__file__).parent
    twofile_path = analytics_path / "2-file"

    # Маппинг версий для обеих папок
    version_mapping = {
        'analytics': {
            ' version_1': 'v1_english',
            'version_2': 'v2_russian',
            'version_3': 'v3_russian_v2'
        },
        '2-file': {
            'v1': 'v1_english',
            'V2': 'v2_russian',
            'v3': 'v3_russian_v2'
        }
    }

    all_metrics = []

    # Загружаем данные из analytics/
    print("\n📂 Обработка папки analytics/")
    print("-" * 80)
    for version_folder, version_label in version_mapping['analytics'].items():
        version_path = analytics_path / version_folder
        if not version_path.exists():
            print(f"⚠️  Версия {version_folder} не найдена, пропускаем")
            continue

        df = load_ragchecker_metrics(version_path, "analytics")
        if not df.empty:
            df['version_label'] = version_label
            all_metrics.append(df)
            print(f"✓ {version_folder} ({version_label}): {len(df)} записей")
            print(f"   Лекции: {df['lecture_file'].unique().tolist()}")

    # Загружаем данные из 2-file/
    print("\n📂 Обработка папки 2-file/")
    print("-" * 80)
    for version_folder, version_label in version_mapping['2-file'].items():
        version_path = twofile_path / version_folder
        if not version_path.exists():
            print(f"⚠️  Версия {version_folder} не найдена, пропускаем")
            continue

        df = load_ragchecker_metrics(version_path, "2-file")
        if not df.empty:
            df['version_label'] = version_label
            all_metrics.append(df)
            print(f"✓ {version_folder} ({version_label}): {len(df)} записей")
            print(f"   Лекции: {df['lecture_file'].unique().tolist()}")

    if not all_metrics:
        print("\n❌ Не найдено данных для анализа!")
        return

    # Объединяем все данные
    combined_df = pd.concat(all_metrics, ignore_index=True)

    print("\n" + "=" * 80)
    print("📊 ОБЩАЯ СТАТИСТИКА")
    print("=" * 80)
    print(f"Всего записей: {len(combined_df)}")
    print(f"Файлов лекций: {combined_df['lecture_file'].nunique()}")
    print(f"Версий: {combined_df['version_label'].nunique()}")
    print(f"Моделей: {combined_df['model_name'].nunique()}")
    print(f"Промптов: {combined_df['prompt_id'].nunique()}")

    print("\nФайлы лекций:")
    for lecture in combined_df['lecture_file'].unique():
        count = len(combined_df[combined_df['lecture_file'] == lecture])
        print(f"  - {lecture}: {count} записей")

    # Создаем Excel отчет (в корне проекта, на уровень выше analytics)
    project_root = analytics_path.parent
    output_file = project_root / "lecture_comparison_analysis.xlsx"
    output_csv = project_root / "lecture_comparison_analysis.csv"
    print(f"\n💾 Создание отчета: {output_file}")

    # Сохраняем CSV файл с полными данными
    combined_df.to_csv(output_csv, index=False)
    print(f"💾 Сохранен CSV файл: {output_csv}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Лист 1: Все данные
        combined_df.to_excel(writer, sheet_name='Все данные', index=False)

        # Лист 2: Сравнение по файлам лекций
        lecture_comparison = combined_df.groupby(['lecture_file', 'version_label']).agg({
            'precision': ['mean', 'std', 'count'],
            'recall': ['mean', 'std', 'count'],
            'f1': ['mean', 'std', 'count'],
            'hallucination': ['mean', 'std', 'count'],
            'faithfulness': ['mean', 'std', 'count'],
        }).round(2)
        lecture_comparison.to_excel(writer, sheet_name='Сравнение по лекциям')

        # Лист 3: Сравнение по моделям (все лекции)
        model_comparison = combined_df.groupby(['model_name', 'lecture_file']).agg({
            'precision': 'mean',
            'recall': 'mean',
            'f1': 'mean',
            'hallucination': 'mean',
            'faithfulness': 'mean',
        }).round(2)
        model_comparison.to_excel(writer, sheet_name='Модели по лекциям')

        # Лист 4: Сравнение по версиям (все лекции)
        version_comparison = combined_df.groupby(['version_label', 'lecture_file']).agg({
            'precision': 'mean',
            'recall': 'mean',
            'f1': 'mean',
            'hallucination': 'mean',
            'faithfulness': 'mean',
        }).round(2)
        version_comparison.to_excel(writer, sheet_name='Версии по лекциям')

        # Лист 5: Лучшие комбинации для каждой лекции
        best_configs = combined_df.loc[combined_df.groupby('lecture_file')['f1'].idxmax()]
        best_configs[['lecture_file', 'model_name', 'version_label', 'prompt_id', 'f1', 'precision', 'recall']].to_excel(
            writer, sheet_name='Лучшие конфигурации', index=False
        )

        # Листы 6+: Матрицы F1 для каждой лекции × версии
        for lecture in combined_df['lecture_file'].unique():
            for version in combined_df['version_label'].unique():
                subset = combined_df[
                    (combined_df['lecture_file'] == lecture) &
                    (combined_df['version_label'] == version)
                ]

                if subset.empty:
                    continue

                pivot = subset.pivot_table(
                    index='model_name',
                    columns='prompt_id',
                    values='f1',
                    aggfunc='mean'
                )

                if not pivot.empty:
                    # Укорачиваем имя лекции для имени листа
                    lecture_short = lecture[:20] if len(lecture) > 20 else lecture
                    sheet_name = f"F1 {lecture_short[:15]}_{version[:3]}"
                    # Excel не любит некоторые символы в именах листов
                    sheet_name = sheet_name.replace(':', '_').replace('/', '_').replace('\\', '_')
                    pivot.to_excel(writer, sheet_name=sheet_name[:31])  # Макс 31 символ

    # Применяем условное форматирование
    print("🎨 Применение условного форматирования...")
    wb = load_workbook(output_file)

    # Форматируем основные листы
    for sheet_name in ['Все данные', 'Сравнение по лекциям', 'Модели по лекциям', 'Версии по лекциям']:
        if sheet_name in wb.sheetnames:
            style_worksheet(wb[sheet_name])

    # Форматируем матрицы F1
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('F1 '):
            ws = wb[sheet_name]
            if ws.max_row > 1 and ws.max_column > 1:
                apply_conditional_formatting(ws, 2, 2, ws.max_row, ws.max_column, 'higher_is_better')

    wb.save(output_file)

    print(f"\n✅ Отчеты успешно созданы:")
    print(f"   📊 Excel: {output_file}")
    print(f"   📄 CSV:   {output_csv}")

    # Выводим краткую статистику
    print("\n" + "=" * 80)
    print("📊 КРАТКИЕ РЕЗУЛЬТАТЫ")
    print("=" * 80)

    summary = combined_df.groupby(['lecture_file', 'version_label']).agg({
        'f1': 'mean',
        'precision': 'mean',
        'recall': 'mean',
        'hallucination': 'mean'
    }).round(2)

    print("\nСредние метрики по лекциям и версиям:")
    print(summary)

    print("\n" + "=" * 80)
    print("АНАЛИЗ ЗАВЕРШЕН")
    print("=" * 80)


if __name__ == "__main__":
    main()
